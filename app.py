# app.py
import os, json
import sqlite3
from datetime import datetime, timedelta
from flask import Flask, g, render_template, request, redirect, url_for, flash, jsonify, send_file, session
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from flask_bcrypt import Bcrypt

app = Flask(__name__)
app.config['SECRET_KEY'] = os.urandom(24)
app.config['DATABASE'] = os.path.join(app.instance_path, 'database.db')

os.makedirs(app.instance_path, exist_ok=True)

TRANSLATIONS = {
    'en': {
        'dashboard': 'Dashboard',
        'manage_projects': 'Manage Projects',
        'hr_reports': 'HR Reports',
        'change_password': 'Change Password',
        'logout': 'Logout',
        'welcome': 'Welcome',
        'projects': 'Projects',
        'employees': 'Employees',
        'tasks': 'Tasks',
        'add_project': 'Add New Project',
        'project_name': 'Project Name',
        'client': 'Client',
        'start_date': 'Start Date',
        'end_date': 'End Date',
        'actions': 'Actions',
        'edit': 'Edit',
        'delete': 'Delete',
        'allocated_hours': 'Allocated Hours',
        'worked_hours': 'Worked Hours',
        'remaining_hours': 'Remaining Hours',
        'budget_status': 'Budget Status',
        'over_budget': 'Over Budget',
        'near_limit': 'Near Limit',
        'within_limit': 'Within Limit',
        'unlimited': 'Unlimited',
        'status': 'Status',
        'category': 'Category',
        'add_task': 'Add Task',
        'task_title': 'Task Title',
        'description': 'Description',
        'approval_status': 'Approval Status',
        'submit_approval': 'Submit for Approval',
        'pending_approvals': 'Pending Approvals',
        'approve': 'Approve',
        'reject': 'Reject',
        'comments': 'Comments',
        'rejection_comments': 'Rejection Comments',
        'language': 'Language',
        'save': 'Save',
        'cancel': 'Cancel',
        'top_employees': 'Top Employees (Worked Hours)',
        'project_time_distribution': 'Project Time Distribution',
        'weekly_activity': 'Weekly Activity (Hours)',
        'print_report': 'Print PDF Report',
        'filter': 'Filter',
        'all': 'All',
        'assign_employees': 'Assign Employees',
        'select_all': 'Select All',
        'subordinates': 'Subordinates',
        'subordinate_tasks': 'Subordinate Tasks',
        'add_subordinate_task': 'Assign Task to Subordinate',
        'assigned_to': 'Assigned To',
        'view_details': 'View Details',
        'client_name': 'Client Name',
        'duration': 'Duration',
        'date': 'Date',
        'unlimited_hours': 'Unlimited hours',
        'hrs_remaining': 'hrs remaining',
        'project_details': 'Project Details',
        'employee_details': 'Employee Details',
        'allocated': 'Allocated',
        'worked': 'Worked',
        'remaining': 'Remaining',
        'add_employee': 'Add Employee',
        'full_name': 'Full Name',
        'username': 'Username',
        'position': 'Position',
        'role': 'Role',
        'manager': 'Manager',
        'assign_managers': 'Assign Managers',
        'existing_employees': 'Existing Employees',
        'existing_projects': 'Existing Projects',
        'no_projects_found': 'No projects found in the system.',
        'no_employees_found': 'No employees found in the system.',
        'add_new_task': 'Add New Task',
        'select_project': 'Select Project',
        'select_category': 'Select Category',
        'select_employee': 'Select Employee',
        'subordinate': 'Subordinate',
        'no_tasks_found': 'No tasks found.',
        'date_range': 'Date Range',
        'from': 'From',
        'to': 'To',
        'export_excel': 'Export Excel',
        'status_order': 'Status Order',
        'search': 'Search',
        'submit': 'Submit',
        'timesheet': 'Timesheet',
        'submit_timesheet_desc': 'Submit all finished tasks for approval',
        'submit_finished_tasks': 'Submit Finished Tasks',
        'my_tasks': 'My Tasks',
        'assigned_tasks': 'Assigned Tasks',
        'running': 'Running',
        'pause': 'Paused',
        'finish': 'Finished',
        'not_start': 'Not Started',
        'timesheet_status': 'Timesheet Status',
        'no_pending_approvals': 'No pending approvals.',
        'comments_optional': 'Comments (Optional)'
    },
    'ar': {
        'dashboard': 'لوحة التحكم',
        'manage_projects': 'إدارة المشاريع',
        'hr_reports': 'تقارير الموارد البشرية',
        'change_password': 'تغيير كلمة المرور',
        'logout': 'تسجيل الخروج',
        'welcome': 'مرحباً',
        'projects': 'المشاريع',
        'employees': 'الموظفين',
        'tasks': 'المهام',
        'add_project': 'إضافة مشروع جديد',
        'project_name': 'اسم المشروع',
        'client': 'العميل',
        'start_date': 'تاريخ البدء',
        'end_date': 'تاريخ الانتهاء',
        'actions': 'الإجراءات',
        'edit': 'تعديل',
        'delete': 'حذف',
        'allocated_hours': 'الساعات المخصصة',
        'worked_hours': 'الساعات الفعلية',
        'remaining_hours': 'الساعات المتبقية',
        'budget_status': 'حالة الميزانية',
        'over_budget': 'تجاوز الميزانية',
        'near_limit': 'قريب من الحد',
        'within_limit': 'ضمن الحد',
        'unlimited': 'غير محدود',
        'status': 'الحالة',
        'category': 'الفئة',
        'add_task': 'إضافة مهمة',
        'task_title': 'عنوان المهمة',
        'description': 'الوصف',
        'approval_status': 'حالة الموافقة',
        'submit_approval': 'تقديم للموافقة',
        'pending_approvals': 'موافقات معلقة',
        'approve': 'موافقة',
        'reject': 'رفض',
        'comments': 'الملاحظات',
        'rejection_comments': 'ملاحظات الرفض',
        'language': 'اللغة',
        'save': 'حفظ',
        'cancel': 'إلغاء',
        'top_employees': 'الموظفون الأكثر عملاً (بالساعات)',
        'project_time_distribution': 'توزيع الوقت على المشاريع',
        'weekly_activity': 'النشاط الأسبوعي (بالساعات)',
        'print_report': 'طباعة تقرير PDF',
        'filter': 'تصفية',
        'all': 'الكل',
        'assign_employees': 'تعيين الموظفين',
        'select_all': 'تحديد الكل',
        'subordinates': 'المرؤوسين',
        'subordinate_tasks': 'مهام المرؤوسين',
        'add_subordinate_task': 'إسناد مهمة لمرؤوس',
        'assigned_to': 'مُسند إلى',
        'view_details': 'عرض التفاصيل',
        'client_name': 'اسم العميل',
        'duration': 'المدة',
        'date': 'التاريخ',
        'unlimited_hours': 'ساعات غير محدودة',
        'hrs_remaining': 'ساعة متبقية',
        'project_details': 'تفاصيل المشروع',
        'employee_details': 'تفاصيل الموظف',
        'allocated': 'المخصص',
        'worked': 'الفعلي',
        'remaining': 'المتبقي',
        'add_employee': 'إضافة موظف',
        'full_name': 'الاسم الكامل',
        'username': 'اسم المستخدم',
        'position': 'المسمى الوظيفي',
        'role': 'الدور',
        'manager': 'المدير',
        'assign_managers': 'تعيين المدراء',
        'existing_employees': 'الموظفون الحاليون',
        'existing_projects': 'المشاريع الحالية',
        'no_projects_found': 'لا توجد مشاريع في النظام.',
        'no_employees_found': 'لا يوجد موظفون في النظام.',
        'add_new_task': 'إضافة مهمة جديدة',
        'select_project': 'اختر المشروع',
        'select_category': 'اختر الفئة',
        'select_employee': 'اختر الموظف',
        'subordinate': 'المرؤوس',
        'no_tasks_found': 'لا توجد مهام.',
        'date_range': 'النطاق الزمني',
        'from': 'من',
        'to': 'إلى',
        'export_excel': 'تصدير Excel',
        'status_order': 'ترتيب الحالة',
        'search': 'بحث',
        'submit': 'إرسال',
        'timesheet': 'جدول الساعات',
        'submit_timesheet_desc': 'إرسال جميع المهام المنتهية للمراجعة والموافقة',
        'submit_finished_tasks': 'إرسال المهام المنتهية',
        'my_tasks': 'مهامي',
        'assigned_tasks': 'المهام المُسندة',
        'running': 'قيد التشغيل',
        'pause': 'متوقفة مؤقتاً',
        'finish': 'منتهية',
        'not_start': 'لم تبدأ',
        'timesheet_status': 'حالة جدول الساعات',
        'no_pending_approvals': 'لا توجد موافقات معلقة.',
        'comments_optional': 'ملاحظات (اختياري)'
    }
}

import urllib.parse

def get_user_preferences(user_id):
    db = get_db()
    row = db.execute('SELECT * FROM user_preferences WHERE user_id = ?', (user_id,)).fetchone()
    if not row:
        try:
            db.execute('INSERT INTO user_preferences (user_id, lang, filters) VALUES (?, ?, ?)', (user_id, 'ar', '{}'))
            db.commit()
        except sqlite3.IntegrityError:
            pass
        return {'user_id': user_id, 'lang': 'ar', 'filters': {}}
    
    filters = {}
    if row['filters']:
        try:
            filters = json.loads(row['filters'])
        except Exception:
            filters = {}
    return {
        'user_id': row['user_id'],
        'lang': row['lang'] or 'ar',
        'filters': filters
    }

def save_user_preferences(user_id, lang=None, filters=None):
    db = get_db()
    pref = get_user_preferences(user_id)
    new_lang = lang if lang is not None else pref['lang']
    new_filters = pref['filters']
    if filters is not None:
        for page_key, page_val in filters.items():
            new_filters[page_key] = page_val
            
    db.execute('''
        INSERT OR REPLACE INTO user_preferences (user_id, lang, filters)
        VALUES (?, ?, ?)
    ''', (user_id, new_lang, json.dumps(new_filters)))
    db.commit()

@app.before_request
def load_user_preferences_to_session():
    if request.endpoint == 'static':
        return
    if current_user and current_user.is_authenticated:
        pref = get_user_preferences(current_user.id)
        if session.get('lang') != pref['lang']:
            session['lang'] = pref['lang']

@app.context_processor
def utility_processor():
    def translate(key):
        lang = session.get('lang', 'ar')
        return TRANSLATIONS.get(lang, {}).get(key, key)
    return dict(_=translate, current_lang=session.get('lang', 'ar'))

@app.route('/set_language/<lang>')
def set_language(lang):
    if lang in ['ar', 'en']:
        session['lang'] = lang
        if current_user and current_user.is_authenticated:
            save_user_preferences(current_user.id, lang=lang)
    return redirect(request.referrer or url_for('index'))

login_manager = LoginManager()
login_manager.login_view = 'login'
login_manager.init_app(app)

bcrypt = Bcrypt(app)
import random, re
from faker import Faker
@app.template_filter('datetimeformat')
def datetimeformat(value):
    if not value:
        return ""
    try:
        if len(value) >= 16 and value[10] == ' ' and value[13] == ':':
            return value[:16]
        if 'T' in value:
            parts = value.split('T')
            date_part = parts[0]
            time_part = parts[1].split('.')[0]
            time_part = ':'.join(time_part.split(':')[:2])
            return f"{date_part} {time_part}"
        return value
    except Exception:
        return value

@app.template_filter('format_duration')
def format_duration(task):
    try:
        duration = task['running_duration']
        if duration is None:
            duration = 0
    except Exception:
        duration = 0
        
    try:
        status = task['status']
        last_started_at = task['last_started_at']
    except Exception:
        status = None
        last_started_at = None
        
    if status == 'Running' and last_started_at:
        try:
            started_dt = datetime.fromisoformat(last_started_at)
            elapsed = (datetime.utcnow() - started_dt).total_seconds()
            duration += max(0, int(elapsed))
        except Exception:
            pass
            
    hours = duration // 3600
    minutes = (duration % 3600) // 60
    seconds = duration % 60
    
    parts = []
    if hours > 0:
        parts.append(f"{hours}h")
    if minutes > 0:
        parts.append(f"{minutes}m")
    if seconds > 0 or not parts:
        parts.append(f"{seconds}s")
        
    return " ".join(parts)

@app.template_filter('total_seconds')
def total_seconds(task):
    try:
        duration = task['running_duration']
        if duration is None:
            duration = 0
    except Exception:
        duration = 0
        
    try:
        status = task['status']
        last_started_at = task['last_started_at']
    except Exception:
        status = None
        last_started_at = None
        
    if status == 'Running' and last_started_at:
        try:
            started_dt = datetime.fromisoformat(last_started_at)
            elapsed = (datetime.utcnow() - started_dt).total_seconds()
            duration += max(0, int(elapsed))
        except Exception:
            pass
    return duration

# ---------- Database helpers ----------

def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(app.config['DATABASE'])
        g.db.row_factory = sqlite3.Row
    return g.db

@app.teardown_appcontext
def close_db(e=None):
    db = g.pop('db', None)
    if db is not None:
        db.close()

def init_db():
    db = get_db()
    # Users table: roles are HR or Employee
    db.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            full_name TEXT NOT NULL,
            username TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL,
            position TEXT
        )
    ''')
    # Join table for many-to-many employee-manager relationship (both roles are Employee)
    db.execute('''
        CREATE TABLE IF NOT EXISTS employee_managers (
            employee_id INTEGER NOT NULL,
            manager_id INTEGER NOT NULL,
            FOREIGN KEY(employee_id) REFERENCES users(id),
            FOREIGN KEY(manager_id) REFERENCES users(id),
            PRIMARY KEY (employee_id, manager_id)
        )
    ''')
    # Join table for many-to-many employee-project relationship
    db.execute('''
        CREATE TABLE IF NOT EXISTS employee_projects (
            employee_id INTEGER NOT NULL,
            project_id INTEGER NOT NULL,
            allocated_hours INTEGER DEFAULT 0,
            FOREIGN KEY(employee_id) REFERENCES users(id) ON DELETE CASCADE,
            FOREIGN KEY(project_id) REFERENCES projects(id) ON DELETE CASCADE,
            PRIMARY KEY (employee_id, project_id)
        )
    ''')
    try:
        db.execute('ALTER TABLE employee_projects ADD COLUMN allocated_hours INTEGER DEFAULT 0')
        db.commit()
    except sqlite3.OperationalError:
        pass
    # Projects table
    db.execute('''
        CREATE TABLE IF NOT EXISTS projects (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            client TEXT NOT NULL,
            start_date TEXT,
            end_date TEXT
        )
    ''')
    # Tasks table: directly assigned/owned by the employee, created by creator_id
    db.execute('''
        CREATE TABLE IF NOT EXISTS tasks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            employee_id INTEGER NOT NULL,
            project_id INTEGER NOT NULL,
            title TEXT NOT NULL,
            description TEXT,
            status TEXT NOT NULL,
            created_at TEXT NOT NULL,
            running_duration INTEGER DEFAULT 0,
            last_started_at TEXT,
            creator_id INTEGER NOT NULL,
            FOREIGN KEY(employee_id) REFERENCES users(id),
            FOREIGN KEY(creator_id) REFERENCES users(id),
            FOREIGN KEY(project_id) REFERENCES projects(id)
        )
    ''')
    try:
        db.execute('ALTER TABLE tasks ADD COLUMN category TEXT')
    except sqlite3.OperationalError:
        pass
    try:
        db.execute('ALTER TABLE tasks ADD COLUMN approval_status TEXT DEFAULT "Draft"')
    except sqlite3.OperationalError:
        pass
    try:
        db.execute('ALTER TABLE tasks ADD COLUMN approval_comments TEXT')
    except sqlite3.OperationalError:
        pass
    try:
        db.execute('ALTER TABLE tasks ADD COLUMN approver_id INTEGER')
    except sqlite3.OperationalError:
        pass
        
    db.execute('''
        CREATE TABLE IF NOT EXISTS user_preferences (
            user_id INTEGER PRIMARY KEY,
            lang TEXT,
            filters TEXT,
            FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
        )
    ''')
    db.commit()

def seed_data():
    db = get_db()
    if db.execute('SELECT 1 FROM users WHERE username = ?', ('hr',)).fetchone():
        return
    
    # HR User
    hr_pwd = bcrypt.generate_password_hash('123').decode('utf-8')
    db.execute('INSERT INTO users (full_name, username, password_hash, role, position) VALUES (?,?,?,?,?)',
               ('HR User', 'hr', hr_pwd, 'HR', 'HR Manager'))
    
    # Projects
    db.execute('INSERT INTO projects (name, client, start_date, end_date) VALUES (?,?,?,?)', ('Project Alpha', 'Company A', '2026-01-01', '2026-12-31'))
    db.execute('INSERT INTO projects (name, client, start_date, end_date) VALUES (?,?,?,?)', ('Project Beta', 'Company B', '2026-06-01', '2026-11-30'))
    proj_a = db.execute('SELECT id FROM projects WHERE name = ?', ('Project Alpha',)).fetchone()
    proj_b = db.execute('SELECT id FROM projects WHERE name = ?', ('Project Beta',)).fetchone()
    
    # Employees (who can also be managers)
    ahmed_pwd = bcrypt.generate_password_hash('123').decode('utf-8')
    db.execute('INSERT INTO users (full_name, username, password_hash, role, position) VALUES (?,?,?,?,?)',
               ('Ahmed', 'ahmed', ahmed_pwd, 'Employee', 'Software Engineer'))
    
    badr_pwd = bcrypt.generate_password_hash('123').decode('utf-8')
    db.execute('INSERT INTO users (full_name, username, password_hash, role, position) VALUES (?,?,?,?,?)',
               ('Badr', 'badr', badr_pwd, 'Employee', 'Team Lead'))
    
    # Ahmed has Badr as manager
    ahmed = db.execute('SELECT id FROM users WHERE username = ?', ('ahmed',)).fetchone()
    badr = db.execute('SELECT id FROM users WHERE username = ?', ('badr',)).fetchone()
    if ahmed and badr:
        db.execute('INSERT INTO employee_managers (employee_id, manager_id) VALUES (?,?)', (ahmed['id'], badr['id']))
        
        # Assign Ahmed to Project Alpha
        if proj_a:
            db.execute('INSERT OR IGNORE INTO employee_projects (employee_id, project_id) VALUES (?,?)', (ahmed['id'], proj_a['id']))
            
        # Assign Badr to Project Alpha and Beta
        if proj_a:
            db.execute('INSERT OR IGNORE INTO employee_projects (employee_id, project_id) VALUES (?,?)', (badr['id'], proj_a['id']))
        if proj_b:
            db.execute('INSERT OR IGNORE INTO employee_projects (employee_id, project_id) VALUES (?,?)', (badr['id'], proj_b['id']))
            
        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        db.execute('INSERT INTO tasks (employee_id, project_id, title, description, status, created_at, running_duration, creator_id) VALUES (?,?,?,?,?,?,?,?)',
                   (ahmed['id'], proj_a['id'] if proj_a else None, 'Task One', 'First sample task', 'Pause', now, 0, ahmed['id']))
        db.execute('INSERT INTO tasks (employee_id, project_id, title, description, status, created_at, running_duration, creator_id) VALUES (?,?,?,?,?,?,?,?)',
                   (ahmed['id'], proj_a['id'] if proj_a else None, 'Task Two', 'Second sample task', 'Finish', now, 7200, badr['id']))
        
    db.commit()

def seed_mock_data():
    """Populate the SQLite database with mock projects, employees, and tasks."""
    db = get_db()
    # Clear existing data
    db.execute('DELETE FROM tasks')
    db.execute('DELETE FROM employee_managers')
    db.execute('DELETE FROM employee_projects')
    db.execute('DELETE FROM users WHERE role = "Employee"')
    db.execute('DELETE FROM projects')
    db.commit()

    fake = Faker()
    # Create projects
    project_ids = []
    for _ in range(30):
        name = fake.company() + " Project"
        client = fake.company()
        start_date = fake.date_between(start_date='-180d')
        end_date = fake.date_between(start_date=start_date, end_date='+180d')
        cursor = db.execute('INSERT INTO projects (name, client, start_date, end_date) VALUES (?,?,?,?)',
                       (name, client, start_date.isoformat(), end_date.isoformat()))
        project_ids.append(cursor.lastrowid)
    db.commit()

    # Create employees
    employee_ids = []
    default_pwd = bcrypt.generate_password_hash('password123').decode('utf-8')
    for _ in range(50):
        full_name = fake.name()
        username = fake.user_name()
        position = fake.job()
        cursor = db.execute('INSERT INTO users (full_name, username, password_hash, role, position) VALUES (?,?,?,?,?)',
                    (full_name, username, default_pwd, 'Employee', position))
        employee_ids.append(cursor.lastrowid)
    db.commit()

    # Assign managers (up to 2 per employee)
    for emp_id in employee_ids:
        possible_mgrs = [mid for mid in employee_ids if mid != emp_id]
        mgrs = random.sample(possible_mgrs, k=random.randint(0, 2))
        for mgr in mgrs:
            db.execute('INSERT INTO employee_managers (employee_id, manager_id) VALUES (?,?)',
                       (emp_id, mgr))
    db.commit()

    # Assign projects to employees (1 to 5 random projects) and set allocated hours
    for emp_id in employee_ids:
        emp_projects = random.sample(project_ids, k=random.randint(1, 5))
        for p_id in emp_projects:
            allocated = random.choice([0, 10, 20, 45, 75, 120, 160]) # 0 means unlimited
            db.execute('INSERT INTO employee_projects (employee_id, project_id, allocated_hours) VALUES (?,?,?)', (emp_id, p_id, allocated))
    db.commit()

    # Create tasks
    now = datetime.utcnow()
    total_tasks_created = 0
    for emp_id in employee_ids:
        # Fetch assigned projects and their allocations
        rows = db.execute('SELECT project_id, allocated_hours FROM employee_projects WHERE employee_id = ?', (emp_id,)).fetchall()
        for row in rows:
            p_id = row['project_id']
            allocated = row['allocated_hours']
            
            # Determine target total seconds worked on this project
            if allocated > 0:
                # 30% chance to exceed the allocated hours (generating negative balance)
                if random.random() < 0.3:
                    target_hours = allocated * random.uniform(1.05, 1.30)
                else:
                    # 70% chance to stay within limit (generating positive balance)
                    target_hours = allocated * random.uniform(0.15, 0.85)
                target_seconds = int(target_hours * 3600)
            else:
                # Unlimited projects get random hours between 5 and 60
                target_seconds = random.randint(5 * 3600, 60 * 3600)
                
            # Distribute target_seconds among several tasks
            remaining_seconds = target_seconds
            while remaining_seconds > 0:
                # Each task takes between 1 and 8 hours (3600 and 28800 seconds)
                duration = random.randint(3600, min(28800, remaining_seconds))
                if remaining_seconds - duration < 3600:
                    # dump the small tail remainder here
                    duration = remaining_seconds
                
                remaining_seconds -= duration
                
                # Distribute tasks over the last 90 days
                days_ago = random.randint(1, 90)
                created_at = now - timedelta(days=days_ago,
                                             hours=random.randint(0, 23),
                                             minutes=random.randint(0, 59))
                created_str = created_at.strftime('%Y-%m-%d %H:%M:%S')
                
                status = random.choice(['Finish', 'Pause'])
                title = f"{fake.catch_phrase()} Task"
                desc = fake.sentence()
                
                db.execute('INSERT INTO tasks (employee_id, project_id, title, description, status, created_at, running_duration, creator_id) VALUES (?,?,?,?,?,?,?,?)',
                           (emp_id, p_id, title, desc, status, created_str, duration, emp_id))
                total_tasks_created += 1
    db.commit()
    return {'projects': len(project_ids), 'employees': len(employee_ids), 'tasks_generated': total_tasks_created}

# ---------- Manager helpers ----------

def get_manager_ids(employee_id):
    db = get_db()
    rows = db.execute('SELECT manager_id FROM employee_managers WHERE employee_id = ?', (employee_id,)).fetchall()
    return [r['manager_id'] for r in rows]

def set_manager_ids(employee_id, manager_ids):
    db = get_db()
    db.execute('DELETE FROM employee_managers WHERE employee_id = ?', (employee_id,))
    for mid in manager_ids:
        db.execute('INSERT INTO employee_managers (employee_id, manager_id) VALUES (?,?)', (employee_id, mid))
    db.commit()

# ---------- User model ----------
class User(UserMixin):
    def __init__(self, id_, username, role, full_name=None):
        self.id = id_
        self.username = username
        self.role = role
        self.full_name = full_name

    @staticmethod
    def get(user_id):
        row = get_db().execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()
        if row:
            return User(row['id'], row['username'], row['role'], row['full_name'])
        return None

    @staticmethod
    def find_by_username(username):
        row = get_db().execute('SELECT * FROM users WHERE username = ?', (username,)).fetchone()
        if row:
            return User(row['id'], row['username'], row['role'], row['full_name'])
        return None

@login_manager.user_loader
def load_user(user_id):
    return User.get(user_id)

# ---------- Role helper ----------
def role_required(*roles):
    def decorator(f):
        @login_required
        def wrapped(*args, **kwargs):
            if current_user.role not in roles:
                flash('Access denied for your role.', 'danger')
                return redirect(url_for('login'))
            return f(*args, **kwargs)
        wrapped.__name__ = f.__name__
        return wrapped
    return decorator

# ---------- Routes ----------
@app.route('/')
def index():
    if current_user.is_authenticated:
        if current_user.role == 'HR':
            return redirect(url_for('hr_dashboard'))
        elif current_user.role == 'Employee':
            return redirect(url_for('employee_dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        row = get_db().execute('SELECT * FROM users WHERE username = ?', (username,)).fetchone()
        if row and bcrypt.check_password_hash(row['password_hash'], password):
            user = User(row['id'], row['username'], row['role'], row['full_name'])
            login_user(user)
            flash('Logged in successfully.', 'success')
            return redirect(url_for('index'))
        flash('Invalid credentials.', 'danger')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Logged out.', 'info')
    return redirect(url_for('login'))

@app.route('/change_password', methods=['GET', 'POST'])
@login_required
def change_password():
    if request.method == 'POST':
        current_password = request.form['current_password']
        new_password = request.form['new_password']
        confirm_password = request.form['confirm_password']
        
        if new_password != confirm_password:
            flash('New passwords do not match.', 'danger')
            return render_template('change_password.html')
            
        db = get_db()
        row = db.execute('SELECT * FROM users WHERE id = ?', (current_user.id,)).fetchone()
        
        if row and bcrypt.check_password_hash(row['password_hash'], current_password):
            new_hash = bcrypt.generate_password_hash(new_password).decode('utf-8')
            db.execute('UPDATE users SET password_hash = ? WHERE id = ?', (new_hash, current_user.id))
            db.commit()
            flash('Password changed successfully.', 'success')
            return redirect(url_for('index'))
        else:
            flash('Incorrect current password.', 'danger')
            
    return render_template('change_password.html')

# ----- HR Dashboard & User Management -----
@app.route('/hr')
@role_required('HR')
def hr_dashboard():
    db = get_db()
    
    if 'reset' in request.args:
        save_user_preferences(current_user.id, filters={'hr_dashboard': {}})
        return redirect(url_for('hr_dashboard'))
        
    if 'filter_applied' in request.args:
        new_filters = {
            'sum_employees': request.args.getlist('sum_employees'),
            'sum_statuses': request.args.getlist('sum_statuses'),
            'summary_created_from': request.args.get('summary_created_from', ''),
            'summary_created_to': request.args.get('summary_created_to', '')
        }
        save_user_preferences(current_user.id, filters={'hr_dashboard': new_filters})
    else:
        pref = get_user_preferences(current_user.id)
        saved = pref.get('filters', {}).get('hr_dashboard', {})
        if saved:
            params = []
            for k, v in saved.items():
                if isinstance(v, list):
                    for item in v:
                        params.append((k, item))
                else:
                    if v:
                        params.append((k, v))
            if params:
                params.append(('filter_applied', '1'))
                return redirect(url_for('hr_dashboard') + '?' + urllib.parse.urlencode(params))
                
    total_employees = db.execute("SELECT COUNT(*) FROM users WHERE role = 'Employee'").fetchone()[0]
    # Fetch employee task summary
    # Prepare filters
    selected_sum_employees = [int(x) for x in request.args.getlist('sum_employees') if x.isdigit()]
    selected_sum_statuses = request.args.getlist('sum_statuses')  # values: Running, Pause, Finish, Not Work
    
    summary_query = '''
        SELECT u.id, u.full_name AS employee_name,
            SUM(CASE WHEN t.status = 'Running' THEN 1 ELSE 0 END) AS running_cnt,
            SUM(CASE WHEN t.status = 'Pause' THEN 1 ELSE 0 END) AS pause_cnt,
            SUM(CASE WHEN t.status = 'Finish' THEN 1 ELSE 0 END) AS finish_cnt
        FROM users u
        LEFT JOIN tasks t ON u.id = t.employee_id
        WHERE u.role = 'Employee'
    '''
    summary_params = []
    if selected_sum_employees:
        placeholders = ','.join('?' for _ in selected_sum_employees)
        summary_query += f" AND u.id IN ({placeholders})"
        summary_params.extend(selected_sum_employees)
    # Date range filters for summary (apply to task creation date)
    created_from = request.args.get('summary_created_from', '').strip()
    created_to = request.args.get('summary_created_to', '').strip()
    if created_from:
        summary_query += " AND t.created_at >= ?"
        summary_params.append(f"{created_from} 00:00")
    if created_to:
        summary_query += " AND t.created_at <= ?"
        summary_params.append(f"{created_to} 23:59")
    summary_query += " GROUP BY u.id"
    employee_summaries = db.execute(summary_query, summary_params).fetchall()
    # Convert to list of dicts with counts
    summaries = []
    for row in employee_summaries:
        total = (row['running_cnt'] or 0) + (row['pause_cnt'] or 0) + (row['finish_cnt'] or 0)
        # Apply status filter if any
        include = True
        if selected_sum_statuses:
            # For each selected status, check if employee has count>0 (except Not Work)
            status_match = False
            for status in selected_sum_statuses:
                if status == 'Running' and (row['running_cnt'] or 0) > 0:
                    status_match = True
                if status == 'Pause' and (row['pause_cnt'] or 0) > 0:
                    status_match = True
                if status == 'Finish' and (row['finish_cnt'] or 0) > 0:
                    status_match = True
                if status == 'Not Work' and (row['running_cnt'] or 0) == 0:
                    status_match = True
            include = status_match
        if include:
            summaries.append({
                'employee_id': row['id'],
                'employee_name': row['employee_name'],
                'running_cnt': row['running_cnt'] or 0,
                'pause_cnt': row['pause_cnt'] or 0,
                'finish_cnt': row['finish_cnt'] or 0,
                'total': total
            })
    # Fetch employee options for filter UI
    employees_list = db.execute("SELECT id, full_name FROM users WHERE role = 'Employee' ORDER BY full_name").fetchall()
    total_projects = db.execute("SELECT COUNT(*) FROM projects").fetchone()[0]
    active_tasks = db.execute("SELECT COUNT(*) FROM tasks WHERE status = 'Running'").fetchone()[0]

    # Fetch project time distribution
    proj_dist = db.execute('''
        SELECT p.name, COALESCE(SUM(t.running_duration), 0) AS total_duration
        FROM projects p
        LEFT JOIN tasks t ON p.id = t.project_id
        GROUP BY p.id
        HAVING total_duration > 0
    ''').fetchall()
    hr_chart_projects_labels = [r['name'] for r in proj_dist]
    hr_chart_projects_data = [round(r['total_duration'] / 3600.0, 2) for r in proj_dist]
    
    # Fetch top 5 employees by hours
    top_emp = db.execute('''
        SELECT u.full_name, COALESCE(SUM(t.running_duration), 0) AS total_duration
        FROM users u
        JOIN tasks t ON u.id = t.employee_id
        WHERE u.role != 'HR'
        GROUP BY u.id
        HAVING total_duration > 0
        ORDER BY total_duration DESC
        LIMIT 5
    ''').fetchall()
    hr_chart_employees_labels = [r['full_name'] for r in top_emp]
    hr_chart_employees_data = [round(r['total_duration'] / 3600.0, 2) for r in top_emp]
    
    # Fetch pending approvals (excluding tasks that do not require approval)
    pending_approvals_raw = db.execute('''
        SELECT t.*, u.full_name AS employee_name, p.name AS project_name, approver.full_name AS approver_name
        FROM tasks t
        JOIN users u ON t.employee_id = u.id
        LEFT JOIN projects p ON t.project_id = p.id
        LEFT JOIN users approver ON t.approver_id = approver.id
        WHERE t.approval_status = 'Submitted' AND t.approver_id IS NOT NULL
        ORDER BY t.created_at DESC
    ''').fetchall()
    
    pending_approvals = []
    for row in pending_approvals_raw:
        task_dict = dict(row)
        try:
            created_dt = None
            try:
                created_dt = datetime.strptime(row['created_at'], '%Y-%m-%d %H:%M:%S')
            except ValueError:
                created_dt = datetime.strptime(row['created_at'], '%Y-%m-%d %H:%M')
            
            delta = datetime.now() - created_dt
            if delta.days > 0:
                pending_since = f"{delta.days}d" if session.get('lang') != 'ar' else f"منذ {delta.days} يوم"
            else:
                hours = delta.seconds // 3600
                minutes = (delta.seconds % 3600) // 60
                if hours > 0:
                    pending_since = f"{hours}h {minutes}m" if session.get('lang') != 'ar' else f"منذ {hours} ساعة و {minutes} د"
                else:
                    pending_since = f"{minutes}m" if session.get('lang') != 'ar' else f"منذ {minutes} د"
        except Exception:
            pending_since = "–"
        task_dict['pending_since'] = pending_since
        pending_approvals.append(task_dict)

    return render_template(
        'hr.html',
        total_employees=total_employees,
        total_projects=total_projects,
        active_tasks=active_tasks,
        summaries=summaries,
        employees_list=employees_list,
        selected_sum_employees=selected_sum_employees,
        selected_sum_statuses=selected_sum_statuses,
        hr_chart_projects_labels=hr_chart_projects_labels,
        hr_chart_projects_data=hr_chart_projects_data,
        hr_chart_employees_labels=hr_chart_employees_labels,
        hr_chart_employees_data=hr_chart_employees_data,
        pending_approvals=pending_approvals
    )


@app.route('/hr/tracking')
@role_required('HR')
def hr_tracking():
    db = get_db()
    
    if 'reset' in request.args:
        save_user_preferences(current_user.id, filters={'hr_tracking': {}})
        return redirect(url_for('hr_tracking'))
        
    if 'filter_applied' in request.args:
        new_filters = {
            'employees': request.args.getlist('employees'),
            'projects': request.args.getlist('projects'),
            'statuses': request.args.getlist('statuses'),
            'created_from': request.args.get('created_from', ''),
            'created_to': request.args.get('created_to', ''),
            'creators': request.args.getlist('creators')
        }
        save_user_preferences(current_user.id, filters={'hr_tracking': new_filters})
    else:
        pref = get_user_preferences(current_user.id)
        saved = pref.get('filters', {}).get('hr_tracking', {})
        if saved:
            params = []
            for k, v in saved.items():
                if isinstance(v, list):
                    for item in v:
                        params.append((k, item))
                else:
                    if v:
                        params.append((k, v))
            if params:
                params.append(('filter_applied', '1'))
                return redirect(url_for('hr_tracking') + '?' + urllib.parse.urlencode(params))
                
    # Get filter inputs
    selected_employees = [int(x) for x in request.args.getlist('employees') if x.isdigit()]
    selected_projects = []
    for x in request.args.getlist('projects'):
        if x == 'none':
            selected_projects.append('none')
        elif x.isdigit():
            selected_projects.append(int(x))
            
    selected_statuses = request.args.getlist('statuses')
    created_from = request.args.get('created_from', '').strip()
    created_to = request.args.get('created_to', '').strip()
    selected_creators = [int(x) for x in request.args.getlist('creators') if x.isdigit()]
    
    # Build query
    query = '''
        SELECT t.*, u.full_name AS employee_name, creator.full_name AS creator_name, p.name AS project_name
        FROM tasks t
        JOIN users u ON t.employee_id = u.id
        LEFT JOIN users creator ON t.creator_id = creator.id
        LEFT JOIN projects p ON t.project_id = p.id
    '''
    
    conditions = []
    params = []
    
    if selected_employees:
        placeholders = ','.join('?' for _ in selected_employees)
        conditions.append(f"t.employee_id IN ({placeholders})")
        params.extend(selected_employees)
        
    if selected_projects:
        proj_conds = []
        has_none = False
        val_projs = []
        for p in selected_projects:
            if p == 'none':
                has_none = True
            else:
                val_projs.append(p)
        if val_projs:
            placeholders = ','.join('?' for _ in val_projs)
            proj_conds.append(f"t.project_id IN ({placeholders})")
            params.extend(val_projs)
        if has_none:
            proj_conds.append("t.project_id IS NULL")
        if proj_conds:
            conditions.append(f"({' OR '.join(proj_conds)})")
            
    if selected_statuses:
        placeholders = ','.join('?' for _ in selected_statuses)
        conditions.append(f"t.status IN ({placeholders})")
        params.extend(selected_statuses)
        
    if created_from:
        conditions.append("t.created_at >= ?")
        params.append(f"{created_from} 00:00")
        
    if created_to:
        conditions.append("t.created_at <= ?")
        params.append(f"{created_to} 23:59")
        
    if selected_creators:
        placeholders = ','.join('?' for _ in selected_creators)
        conditions.append(f"t.creator_id IN ({placeholders})")
        params.extend(selected_creators)
        
    # Count total tasks first
    count_query = '''
        SELECT COUNT(*)
        FROM tasks t
        JOIN users u ON t.employee_id = u.id
        LEFT JOIN users creator ON t.creator_id = creator.id
        LEFT JOIN projects p ON t.project_id = p.id
    '''
    if conditions:
        count_query += " WHERE " + " AND ".join(conditions)
    total_tasks = db.execute(count_query, params).fetchone()[0]

    # Get page and per_page
    page = request.args.get('page', 1, type=int)
    per_page = 50
    total_pages = (total_tasks + per_page - 1) // per_page if total_tasks > 0 else 1
    page = max(1, min(page, total_pages))

    # Paginate query
    query += " LIMIT ? OFFSET ?"
    tasks_params = list(params) + [per_page, (page - 1) * per_page]
    tasks = db.execute(query, tasks_params).fetchall()
    
    # Helper to construct pagination URLs preserving multi-value filters
    def get_page_url(page_num):
        import urllib.parse
        params_list = []
        for key, values in request.args.lists():
            if key == 'page':
                continue
            for val in values:
                params_list.append((key, val))
        params_list.append(('page', str(page_num)))
        return url_for('hr_tracking') + '?' + urllib.parse.urlencode(params_list)

    # Fetch options for filters
    employees_list = db.execute("SELECT id, full_name, username FROM users WHERE role = 'Employee' ORDER BY full_name").fetchall()
    projects_list = db.execute("SELECT id, name FROM projects ORDER BY name").fetchall()
    creators_list = db.execute("SELECT DISTINCT u.id, u.full_name, u.username FROM users u JOIN tasks t ON t.creator_id = u.id ORDER BY u.full_name").fetchall()
    
    showing_from = (page - 1) * per_page + 1 if total_tasks > 0 else 0
    showing_to = min(page * per_page, total_tasks)
    start_page = max(1, page - 2)
    end_page = min(total_pages, page + 2)
    page_range = list(range(start_page, end_page + 1))

    return render_template(
        'hr_tracking.html', 
        tasks=tasks,
        employees_list=employees_list,
        projects_list=projects_list,
        creators_list=creators_list,
        selected_employees=selected_employees,
        selected_projects=selected_projects,
        selected_statuses=selected_statuses,
        created_from=created_from,
        created_to=created_to,
        selected_creators=selected_creators,
        page=page,
        total_pages=total_pages,
        total_tasks=total_tasks,
        page_url=get_page_url,
        showing_from=showing_from,
        showing_to=showing_to,
        page_range=page_range
    )

@app.route('/hr/employees')
@role_required('HR')
def hr_employees():
    db = get_db()
    users = db.execute('SELECT * FROM users WHERE role != ?', ('HR',)).fetchall()
    return render_template('hr_employees.html', users=users)

@app.route('/hr/projects')
@role_required('HR')
def hr_projects():
    db = get_db()
    projects = db.execute('SELECT * FROM projects').fetchall()
    employees = db.execute('SELECT id, full_name, username FROM users WHERE role != "HR" ORDER BY full_name').fetchall()
    return render_template('hr_projects.html', projects=projects, employees=employees)

@app.route('/hr/employees/<int:user_id>')
@role_required('HR')
def hr_employee_detail(user_id):
    db = get_db()
    employee = db.execute('SELECT * FROM users WHERE id = ? AND role != ?', (user_id, 'HR')).fetchone()
    if not employee:
        flash('Employee not found.', 'danger')
        return redirect(url_for('hr_employees'))
        
    projects = db.execute('''
        SELECT p.id, p.name, ep.allocated_hours,
               COALESCE(SUM(t.running_duration), 0) AS worked_seconds
        FROM projects p
        JOIN employee_projects ep ON p.id = ep.project_id
        LEFT JOIN tasks t ON p.id = t.project_id AND t.employee_id = ep.employee_id
        WHERE ep.employee_id = ?
        GROUP BY p.id
        ORDER BY p.name
    ''', (user_id,)).fetchall()
    
    project_details = []
    for p in projects:
        allocated = p['allocated_hours']
        worked_hours = p['worked_seconds'] / 3600.0
        remaining = allocated - worked_hours if allocated > 0 else 0
        project_details.append({
            'id': p['id'],
            'name': p['name'],
            'allocated_hours': allocated,
            'worked_hours': worked_hours,
            'remaining_hours': remaining
        })
        
    return render_template('hr_employee_detail.html', employee=employee, projects=project_details)

@app.route('/hr/projects/<int:project_id>')
@role_required('HR')
def hr_project_detail(project_id):
    db = get_db()
    project = db.execute('SELECT * FROM projects WHERE id = ?', (project_id,)).fetchone()
    if not project:
        flash('Project not found.', 'danger')
        return redirect(url_for('hr_projects'))
        
    employees = db.execute('''
        SELECT u.id, u.full_name, u.username, ep.allocated_hours,
               COALESCE(SUM(t.running_duration), 0) AS worked_seconds
        FROM users u
        JOIN employee_projects ep ON u.id = ep.employee_id
        LEFT JOIN tasks t ON ep.project_id = t.project_id AND t.employee_id = u.id
        WHERE ep.project_id = ?
        GROUP BY u.id
        ORDER BY u.full_name
    ''', (project_id,)).fetchall()
    
    employee_details = []
    for emp in employees:
        allocated = emp['allocated_hours']
        worked_hours = emp['worked_seconds'] / 3600.0
        remaining = allocated - worked_hours if allocated > 0 else 0
        employee_details.append({
            'id': emp['id'],
            'full_name': emp['full_name'],
            'username': emp['username'],
            'allocated_hours': allocated,
            'worked_hours': worked_hours,
            'remaining_hours': remaining
        })
        
    return render_template('hr_project_detail.html', project=project, employees=employee_details)

@app.route('/hr/projects/add', methods=['POST'])
@role_required('HR')
def add_project():
    name = request.form.get('name')
    client = request.form.get('client')
    start_date = request.form.get('start_date')
    end_date = request.form.get('end_date')
    employee_ids = request.form.getlist('employee_ids')
    if name and client:
        db = get_db()
        try:
            cursor = db.execute('INSERT INTO projects (name, client, start_date, end_date) VALUES (?,?,?,?)',
                       (name, client, start_date or None, end_date or None))
            project_id = cursor.lastrowid
            for emp_id in employee_ids:
                db.execute('INSERT OR IGNORE INTO employee_projects (employee_id, project_id, allocated_hours) VALUES (?,?,0)',
                           (int(emp_id), project_id))
            db.commit()
            flash('Project added successfully.', 'success')
        except sqlite3.IntegrityError:
            flash('Project name already exists.', 'danger')
    else:
        flash('Project name and Client are required.', 'danger')
    return redirect(url_for('hr_projects'))

@app.route('/hr/projects/edit/<int:project_id>', methods=['GET', 'POST'])
@role_required('HR')
def edit_project(project_id):
    db = get_db()
    project = db.execute('SELECT * FROM projects WHERE id = ?', (project_id,)).fetchone()
    if not project:
        flash('Project not found.', 'danger')
        return redirect(url_for('hr_projects'))
    
    if request.method == 'POST':
        name = request.form.get('name')
        client = request.form.get('client')
        start_date = request.form.get('start_date')
        end_date = request.form.get('end_date')
        employee_ids = request.form.getlist('employee_ids')
        if name and client:
            try:
                db.execute('UPDATE projects SET name = ?, client = ?, start_date = ?, end_date = ? WHERE id = ?',
                           (name, client, start_date or None, end_date or None, project_id))
                
                db.execute('DELETE FROM employee_projects WHERE project_id = ?', (project_id,))
                for emp_id in employee_ids:
                    db.execute('INSERT INTO employee_projects (employee_id, project_id, allocated_hours) VALUES (?, ?, 100)',
                               (int(emp_id), project_id))
                db.commit()
                flash('Project updated successfully.', 'success')
                return redirect(url_for('hr_projects'))
            except sqlite3.IntegrityError:
                flash('Project name already exists.', 'danger')
        else:
            flash('Project name and Client are required.', 'danger')
            
    employees = db.execute("SELECT id, full_name, username FROM users WHERE role = 'Employee' ORDER BY full_name").fetchall()
    assigned = db.execute("SELECT employee_id FROM employee_projects WHERE project_id = ?", (project_id,)).fetchall()
    assigned_ids = [r['employee_id'] for r in assigned]
            
    return render_template('edit_project.html', project=project, employees=employees, assigned_ids=assigned_ids)
# ----- HR Reporting -----

from io import BytesIO
import pandas as pd
from flask import send_file, jsonify

def _fetch_tasks_for_employee_report(employee_id, date_from, date_to, project_ids=None, category=None):
    """Return list of dicts with task data for employee report.
    Excludes running tasks and filters by creation date.
    """
    db = get_db()
    params = [employee_id, f"{date_from} 00:00", f"{date_to} 23:59"]
    query = '''
        SELECT t.id, t.created_at, t.running_duration, t.title, t.category, t.approval_status, p.id as project_id, p.name as project_name
        FROM tasks t
        JOIN projects p ON t.project_id = p.id
        WHERE t.employee_id = ?
          AND t.status != 'Running'
          AND datetime(t.created_at) BETWEEN ? AND ?
    '''
    if project_ids:
        placeholders = ','.join('?' for _ in project_ids)
        query += f" AND p.id IN ({placeholders})"
        params.extend(project_ids)
    if category:
        query += " AND t.category = ?"
        params.append(category)
    rows = db.execute(query, params).fetchall()
    return [dict(row) for row in rows]

def _fetch_tasks_for_project_report(project_id, date_from, date_to, employee_ids=None, category=None):
    """Return list of dicts with task data for project report.
    If project_id is None, include all projects.
    """
    db = get_db()
    params = [f"{date_from} 00:00", f"{date_to} 23:59"]
    query = '''
        SELECT t.id, t.created_at, t.running_duration, t.title, t.category, t.approval_status, u.id as employee_id, u.full_name as employee_name, p.id as project_id, p.name as project_name
        FROM tasks t
        JOIN users u ON t.employee_id = u.id
        LEFT JOIN projects p ON t.project_id = p.id
        WHERE t.status != 'Running'
          AND datetime(t.created_at) BETWEEN ? AND ?
    '''
    if project_id is not None:
        query += " AND p.id = ?"
        params.append(project_id)
    if employee_ids:
        placeholders = ','.join('?' for _ in employee_ids)
        query += f" AND u.id IN ({placeholders})"
        params.extend(employee_ids)
    if category:
        query += " AND t.category = ?"
        params.append(category)
    rows = db.execute(query, params).fetchall()
    return [dict(row) for row in rows]

def _apply_excel_formatting(workbook):
    """Apply required formatting to both worksheets using openpyxl.
    """
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    header_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    bold_font = Font(bold=True)
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in workbook.worksheets:
        ws.freeze_panes = ws['A2']
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = bold_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border
                if isinstance(cell.value, (int, float)):
                    cell.number_format = numbers.FORMAT_NUMBER_00
        for col in ws.columns:
            max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
            ws.column_dimensions[col[0].column_letter].width = max_len + 2
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
            if str(row[0].value).lower() == 'total':
                for c in ws[row[0].row]:
                    c.font = bold_font
        last_col = ws.max_column
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=last_col).font = bold_font
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

def _generate_employee_excel(rows, employee_name, date_from, date_to, generated_by):
    """Create Excel workbook for employee report.
    Returns a BytesIO object.
    """
    if not rows:
        out = BytesIO()
        with pd.ExcelWriter(out, engine='openpyxl') as writer:
            pd.DataFrame([{'Message': 'No data found for selected filters.'}]).to_excel(writer, index=False, sheet_name='Details')
            pd.DataFrame([{'Message': 'No data found for selected filters.'}]).to_excel(writer, index=False, sheet_name='Summary')
        out.seek(0)
        return out
    df = pd.DataFrame(rows)
    df['date'] = pd.to_datetime(df['created_at']).dt.date
    df['hours'] = df['running_duration'] / 3600.0
    pivot = pd.pivot_table(df, index='date', columns='project_name', values='hours', aggfunc='sum', fill_value=0)
    pivot = pivot.sort_index()
    pivot['Total'] = pivot.sum(axis=1)
    total_row = pivot.sum(axis=0)
    total_row.name = 'Total'
    details_df = pd.concat([pivot, pd.DataFrame([total_row])])
    # Compute per‑project totals (exclude the 'Total' column and the total row)
    per_project = pivot.iloc[:-1][pivot.columns.drop('Total')].sum()
    # Round to two decimals for display consistency
    per_project = per_project.round(2)
    summary_df = pd.DataFrame({
        'Project': per_project.index,
        'Total Hours': per_project.values
    })
    # Grand total should be the sum of the displayed project totals
    grand_total = per_project.sum().round(2)
    summary_df = pd.concat([summary_df, pd.DataFrame({'Project': ['Grand Total'], 'Total Hours': [grand_total]})], ignore_index=True)
    out = BytesIO()
    with pd.ExcelWriter(out, engine='openpyxl') as writer:
        details_df.to_excel(writer, sheet_name='Details')
        ws = writer.book.create_sheet('Summary')
        info = {
            'Report Type': 'Employee Report',
            'Employee': employee_name,
            'Date From': date_from,
            'Date To': date_to,
            'Generated By': generated_by,
            'Generated On': pd.Timestamp.utcnow().strftime('%Y-%m-%d %H:%M:%S')
        }
        r = 1
        for k, v in info.items():
            ws.cell(row=r, column=1, value=k)
            ws.cell(row=r, column=2, value=v)
            r += 1
        table_start = r + 1
        ws.cell(row=table_start, column=1, value='Project')
        ws.cell(row=table_start, column=2, value='Total Hours')
        for i, rec in enumerate(summary_df.itertuples(index=False), start=1):
            ws.cell(row=table_start + i, column=1, value=rec.Project)
            ws.cell(row=table_start + i, column=2, value=rec[1])
        _apply_excel_formatting(writer.book)
    out.seek(0)
    return out

def _generate_project_excel(rows, project_name, date_from, date_to, generated_by):
    """Create Excel workbook for project report.
    Mirrors employee version but pivots on employee.
    """
    if not rows:
        out = BytesIO()
        with pd.ExcelWriter(out, engine='openpyxl') as writer:
            pd.DataFrame([{'Message': 'No data found for selected filters.'}]).to_excel(writer, index=False, sheet_name='Details')
            pd.DataFrame([{'Message': 'No data found for selected filters.'}]).to_excel(writer, index=False, sheet_name='Summary')
        out.seek(0)
        return out
    df = pd.DataFrame(rows)
    df['date'] = pd.to_datetime(df['created_at']).dt.date
    df['hours'] = df['running_duration'] / 3600.0
    pivot = pd.pivot_table(df, index='date', columns='employee_name', values='hours', aggfunc='sum', fill_value=0)
    pivot = pivot.sort_index()
    pivot['Total'] = pivot.sum(axis=1)
    total_row = pivot.sum(axis=0)
    total_row.name = 'Total'
    details_df = pd.concat([pivot, pd.DataFrame([total_row])])
    # Compute per‑employee totals (exclude the 'Total' column and the total row)
    per_employee = pivot.iloc[:-1][pivot.columns.drop('Total')].sum()
    per_employee = per_employee.round(2)
    summary_df = pd.DataFrame({
        'Employee': per_employee.index,
        'Total Hours': per_employee.values
    })
    # Grand total is sum of displayed employee totals
    grand_total = per_employee.sum().round(2)
    summary_df = pd.concat([summary_df, pd.DataFrame({'Employee': ['Grand Total'], 'Total Hours': [grand_total]})], ignore_index=True)
    summary_df = pd.concat([summary_df, pd.DataFrame({'Employee': ['Grand Total'], 'Total Hours': [total_row['Total']]})], ignore_index=True)
    out = BytesIO()
    with pd.ExcelWriter(out, engine='openpyxl') as writer:
        details_df.to_excel(writer, sheet_name='Details')
        ws = writer.book.create_sheet('Summary')
        info = {
            'Report Type': 'Project Report',
            'Project': project_name if project_name else 'All Projects',
            'Date From': date_from,
            'Date To': date_to,
            'Generated By': generated_by,
            'Generated On': pd.Timestamp.utcnow().strftime('%Y-%m-%d %H:%M:%S')
        }
        r = 1
        for k, v in info.items():
            ws.cell(row=r, column=1, value=k)
            ws.cell(row=r, column=2, value=v)
            r += 1
        table_start = r + 1
        ws.cell(row=table_start, column=1, value='Employee')
        ws.cell(row=table_start, column=2, value='Total Hours')
        for i, rec in enumerate(summary_df.itertuples(index=False), start=1):
            ws.cell(row=table_start + i, column=1, value=rec.Employee)
            ws.cell(row=table_start + i, column=2, value=rec[1])
        _apply_excel_formatting(writer.book)
    out.seek(0)
    return out

# ----- HR Reporting Routes -----

@app.route('/hr/reports')
@role_required('HR')
def hr_reports_page():
    db = get_db()
    employees = db.execute("SELECT id, full_name FROM users WHERE role = 'Employee' ORDER BY full_name").fetchall()
    projects = db.execute("SELECT id, name FROM projects ORDER BY name").fetchall()
    return render_template('hr_reports.html', employees=employees, projects=projects)

@app.route('/hr/export', methods=['POST'])
@role_required('HR')
def hr_export():
    payload = request.get_json()
    if not payload:
        return jsonify({'error': 'Invalid request'}), 400
    report_type = payload.get('type')
    date_from = payload.get('date_from')
    date_to = payload.get('date_to')
    category = payload.get('category')
    if not report_type or not date_from or not date_to:
        return jsonify({'error': 'Missing required parameters'}), 400
    generated_by = current_user.full_name or current_user.username
    if report_type == 'employee':
        employee_id = payload.get('employee_id')
        if not employee_id:
            return jsonify({'error': 'Employee ID required'}), 400
        emp = get_db().execute('SELECT full_name FROM users WHERE id = ?', (employee_id,)).fetchone()
        if not emp:
            return jsonify({'error': 'Employee not found'}), 404
        rows = _fetch_tasks_for_employee_report(employee_id, date_from, date_to, payload.get('project_ids'), category)
        excel_io = _generate_employee_excel(rows, emp['full_name'], date_from, date_to, generated_by)
        sanitized_name = re.sub(r'[^A-Za-z0-9]+', '_', emp['full_name']).strip('_')
        filename = f"{sanitized_name}_{date_from}_{date_to}.xlsx"
    elif report_type == 'project':
        project_id = payload.get('project_id')
        project_name = None
        if project_id:
            proj = get_db().execute('SELECT name FROM projects WHERE id = ?', (project_id,)).fetchone()
            if not proj:
                return jsonify({'error': 'Project not found'}), 404
            project_name = proj['name']
        rows = _fetch_tasks_for_project_report(project_id, date_from, date_to, category=category)
        excel_io = _generate_project_excel(rows, project_name, date_from, date_to, generated_by)
        if project_name:
            sanitized_proj = re.sub(r'[^A-Za-z0-9]+', '_', project_name).strip('_')
        else:
            sanitized_proj = 'all_projects'
        filename = f"{sanitized_proj}_{date_from}_{date_to}.xlsx"
    else:
        return jsonify({'error': 'Invalid report type'}), 400
    return send_file(
        excel_io,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/hr/seed', methods=['POST'])
@role_required('HR')
def hr_seed():
    summary = seed_mock_data()
    return jsonify({'message': 'Database seeded', 'summary': summary}), 200

@app.route('/hr/projects/delete/<int:project_id>', methods=['POST'])
@role_required('HR')
def delete_project(project_id):
    db = get_db()
    db.execute('UPDATE tasks SET project_id = NULL WHERE project_id = ?', (project_id,))
    db.execute('DELETE FROM employee_projects WHERE project_id = ?', (project_id,))
    db.execute('DELETE FROM projects WHERE id = ?', (project_id,))
    db.commit()
    flash('Project deleted.', 'info')
    return redirect(url_for('hr_projects'))

@app.route('/hr/add', methods=['GET', 'POST'])
@role_required('HR')
def add_employee():
    db = get_db()
    if request.method == 'POST':
        full_name = request.form['full_name']
        username = request.form['username']
        password = request.form['password']
        position = request.form.get('position', '')
        manager_usernames = request.form.getlist('managers')
        project_ids = request.form.getlist('projects')
        role = 'Employee'
        pwd_hash = bcrypt.generate_password_hash(password).decode('utf-8')
        
        try:
            db.execute('INSERT INTO users (full_name, username, password_hash, role, position) VALUES (?,?,?,?,?)',
                       (full_name, username, pwd_hash, role, position))
            db.commit()
            
            employee_row = db.execute('SELECT id FROM users WHERE username = ?', (username,)).fetchone()
            if employee_row:
                employee_id = employee_row['id']
                manager_ids = []
                for mgr_name in manager_usernames:
                    mgr_row = db.execute('SELECT id FROM users WHERE username = ?', (mgr_name,)).fetchone()
                    if mgr_row:
                        manager_ids.append(mgr_row['id'])
                set_manager_ids(employee_id, manager_ids)
                
                for p_id in project_ids:
                    hours_val = request.form.get(f'allocated_hours_{p_id}', '0').strip()
                    hours = int(hours_val) if hours_val.isdigit() else 0
                    db.execute('INSERT INTO employee_projects (employee_id, project_id, allocated_hours) VALUES (?,?,?)', (employee_id, int(p_id), hours))
                db.commit()
                
            flash('Employee added.', 'success')
            return redirect(url_for('hr_employees'))
        except sqlite3.IntegrityError:
            flash('Username already exists.', 'danger')
            
    # List all current employees to be chosen as managers
    employees = db.execute("SELECT * FROM users WHERE role = 'Employee'").fetchall()
    projects = db.execute("SELECT * FROM projects ORDER BY name").fetchall()
    return render_template('add_employee.html', employees=employees, projects=projects)

@app.route('/hr/edit/<int:user_id>', methods=['GET', 'POST'])
@role_required('HR')
def edit_employee(user_id):
    db = get_db()
    user = db.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()
    if not user:
        flash('User not found.', 'danger')
        return redirect(url_for('hr_employees'))
    
    if request.method == 'POST':
        full_name = request.form['full_name']
        username = request.form['username']
        position = request.form.get('position', '')
        password = request.form.get('password', '').strip()
        manager_usernames = request.form.getlist('managers')
        project_ids = request.form.getlist('projects')
        
        try:
            if password:
                pwd_hash = bcrypt.generate_password_hash(password).decode('utf-8')
                db.execute('UPDATE users SET full_name = ?, username = ?, password_hash = ?, position = ? WHERE id = ?',
                           (full_name, username, pwd_hash, position, user_id))
            else:
                db.execute('UPDATE users SET full_name = ?, username = ?, position = ? WHERE id = ?',
                           (full_name, username, position, user_id))
            db.commit()
            
            manager_ids = []
            for mgr_name in manager_usernames:
                mgr_row = db.execute('SELECT id FROM users WHERE username = ?', (mgr_name,)).fetchone()
                if mgr_row:
                    manager_ids.append(mgr_row['id'])
            set_manager_ids(user_id, manager_ids)
            
            db.execute('DELETE FROM employee_projects WHERE employee_id = ?', (user_id,))
            for p_id in project_ids:
                hours_val = request.form.get(f'allocated_hours_{p_id}', '0').strip()
                hours = int(hours_val) if hours_val.isdigit() else 0
                db.execute('INSERT INTO employee_projects (employee_id, project_id, allocated_hours) VALUES (?,?,?)', (user_id, int(p_id), hours))
            db.commit()
            
            flash('Employee updated.', 'success')
            return redirect(url_for('hr_employees'))
        except sqlite3.IntegrityError:
            flash('Username already exists.', 'danger')
            
    # List other employees to be chosen as managers
    employees = db.execute("SELECT * FROM users WHERE role = 'Employee' AND id != ?", (user_id,)).fetchall()
    current_mgr_ids = get_manager_ids(user_id)
    current_managers = []
    for mid in current_mgr_ids:
        r = db.execute('SELECT username FROM users WHERE id = ?', (mid,)).fetchone()
        if r:
            current_managers.append(r['username'])
            
    projects = db.execute("SELECT * FROM projects ORDER BY name").fetchall()
    current_proj_rows = db.execute("SELECT project_id, allocated_hours FROM employee_projects WHERE employee_id = ?", (user_id,)).fetchall()
    current_project_ids = [r['project_id'] for r in current_proj_rows]
    project_allocations = {r['project_id']: r['allocated_hours'] for r in current_proj_rows}
            
    return render_template('edit_employee.html', user=user, employees=employees, current_managers=current_managers, projects=projects, current_project_ids=current_project_ids, project_allocations=project_allocations)

@app.route('/hr/delete/<int:user_id>', methods=['POST'])
@role_required('HR')
def delete_employee(user_id):
    db = get_db()
    db.execute('DELETE FROM employee_managers WHERE employee_id = ? OR manager_id = ?', (user_id, user_id))
    db.execute('DELETE FROM employee_projects WHERE employee_id = ?', (user_id,))
    db.execute('DELETE FROM tasks WHERE employee_id = ?', (user_id,))
    db.execute('DELETE FROM users WHERE id = ?', (user_id,))
    db.commit()
    flash('Employee deleted.', 'info')
    return redirect(url_for('hr_employees'))

# ----- Employee Dashboard & Tasks -----
@app.route('/employee')
@role_required('Employee')
def employee_dashboard():
    db = get_db()
    
    if 'reset' in request.args:
        save_user_preferences(current_user.id, filters={'employee_dashboard': {}})
        return redirect(url_for('employee_dashboard'))
        
    if 'filter_applied' in request.args:
        new_filters = {
            'projects': request.args.getlist('projects'),
            'statuses': request.args.getlist('statuses'),
            'created_from': request.args.get('created_from', ''),
            'created_to': request.args.get('created_to', ''),
            'creators': request.args.getlist('creators'),
            'sub_employees': request.args.getlist('sub_employees'),
            'sub_projects': request.args.getlist('sub_projects'),
            'sub_statuses': request.args.getlist('sub_statuses'),
            'sub_created_from': request.args.get('sub_created_from', ''),
            'sub_created_to': request.args.get('sub_created_to', ''),
            'sub_creators': request.args.getlist('sub_creators')
        }
        save_user_preferences(current_user.id, filters={'employee_dashboard': new_filters})
    else:
        pref = get_user_preferences(current_user.id)
        saved = pref.get('filters', {}).get('employee_dashboard', {})
        if saved:
            params = []
            for k, v in saved.items():
                if isinstance(v, list):
                    for item in v:
                        params.append((k, item))
                else:
                    if v:
                        params.append((k, v))
            if params:
                params.append(('filter_applied', '1'))
                return redirect(url_for('employee_dashboard') + '?' + urllib.parse.urlencode(params))
    
    # ------------------ My Tasks Filters ------------------
    selected_projects = []
    for x in request.args.getlist('projects'):
        if x == 'none':
            selected_projects.append('none')
        elif x.isdigit():
            selected_projects.append(int(x))
            
    selected_statuses = request.args.getlist('statuses')
    created_from = request.args.get('created_from', '').strip()
    created_to = request.args.get('created_to', '').strip()
    selected_creators = [int(x) for x in request.args.getlist('creators') if x.isdigit()]
    
    # ------------------ Subordinate Tasks Filters ------------------
    selected_sub_employees = [int(x) for x in request.args.getlist('sub_employees') if x.isdigit()]
    selected_sub_projects = []
    for x in request.args.getlist('sub_projects'):
        if x == 'none':
            selected_sub_projects.append('none')
        elif x.isdigit():
            selected_sub_projects.append(int(x))
            
    selected_sub_statuses = request.args.getlist('sub_statuses')
    sub_created_from = request.args.get('sub_created_from', '').strip()
    sub_created_to = request.args.get('sub_created_to', '').strip()
    selected_sub_creators = [int(x) for x in request.args.getlist('sub_creators') if x.isdigit()]
    
    # Build query for employee's own tasks
    query = '''
        SELECT t.*, u.full_name AS creator_name, p.name AS project_name
        FROM tasks t
        LEFT JOIN users u ON t.creator_id = u.id
        LEFT JOIN projects p ON t.project_id = p.id
        WHERE t.employee_id = ?
    '''
    params = [current_user.id]
    
    if selected_projects:
        proj_conds = []
        has_none = False
        val_projs = []
        for p in selected_projects:
            if p == 'none':
                has_none = True
            else:
                val_projs.append(p)
        if val_projs:
            placeholders = ','.join('?' for _ in val_projs)
            proj_conds.append(f"t.project_id IN ({placeholders})")
            params.extend(val_projs)
        if has_none:
            proj_conds.append("t.project_id IS NULL")
        if proj_conds:
            query += f" AND ({' OR '.join(proj_conds)})"
            
    if selected_statuses:
        placeholders = ','.join('?' for _ in selected_statuses)
        query += f" AND t.status IN ({placeholders})"
        params.extend(selected_statuses)
        
    if created_from:
        query += " AND t.created_at >= ?"
        params.append(f"{created_from} 00:00")
        
    if created_to:
        query += " AND t.created_at <= ?"
        params.append(f"{created_to} 23:59")
        
    if selected_creators:
        placeholders = ','.join('?' for _ in selected_creators)
        query += f" AND t.creator_id IN ({placeholders})"
        params.extend(selected_creators)
        
    query += " ORDER BY t.created_at DESC"
    
    tasks = db.execute(query, params).fetchall()
    
    # List employees managed by the current employee
    subordinates = db.execute('''
        SELECT u.* FROM users u
        JOIN employee_managers em ON u.id = em.employee_id
        WHERE em.manager_id = ?
    ''', (current_user.id,)).fetchall()
    
    # Build query for subordinate tasks live tracking
    sub_query = '''
        SELECT t.*, u.full_name AS employee_name, creator.full_name AS creator_name, p.name AS project_name
        FROM tasks t
        JOIN users u ON t.employee_id = u.id
        JOIN employee_managers em ON u.id = em.employee_id
        LEFT JOIN users creator ON t.creator_id = creator.id
        LEFT JOIN projects p ON t.project_id = p.id
        WHERE em.manager_id = ?
          AND EXISTS (
              SELECT 1 FROM employee_projects ep1
              WHERE ep1.employee_id = t.employee_id AND ep1.project_id = t.project_id
          )
          AND EXISTS (
              SELECT 1 FROM employee_projects ep2
              WHERE ep2.employee_id = ? AND ep2.project_id = t.project_id
          )
    '''
    sub_params = [current_user.id, current_user.id]
    
    if selected_sub_employees:
        placeholders = ','.join('?' for _ in selected_sub_employees)
        sub_query += f" AND t.employee_id IN ({placeholders})"
        sub_params.extend(selected_sub_employees)
        
    if selected_sub_projects:
        proj_conds = []
        has_none = False
        val_projs = []
        for p in selected_sub_projects:
            if p == 'none':
                has_none = True
            else:
                val_projs.append(p)
        if val_projs:
            placeholders = ','.join('?' for _ in val_projs)
            proj_conds.append(f"t.project_id IN ({placeholders})")
            sub_params.extend(val_projs)
        if has_none:
            proj_conds.append("t.project_id IS NULL")
        if proj_conds:
            sub_query += f" AND ({' OR '.join(proj_conds)})"
            
    if selected_sub_statuses:
        placeholders = ','.join('?' for _ in selected_sub_statuses)
        sub_query += f" AND t.status IN ({placeholders})"
        sub_params.extend(selected_sub_statuses)
        
    if sub_created_from:
        sub_query += " AND t.created_at >= ?"
        sub_params.append(f"{sub_created_from} 00:00")
        
    if sub_created_to:
        sub_query += " AND t.created_at <= ?"
        sub_params.append(f"{sub_created_to} 23:59")
        
    if selected_sub_creators:
        placeholders = ','.join('?' for _ in selected_sub_creators)
        sub_query += f" AND t.creator_id IN ({placeholders})"
        sub_params.extend(selected_sub_creators)
        
    sub_query += " ORDER BY t.created_at DESC"
    
    subordinate_tasks = db.execute(sub_query, sub_params).fetchall()
    
    # Fetch options for filters specifically for this employee's tasks (restricted to assigned projects)
    projects_list = db.execute('''
        SELECT p.id, p.name FROM projects p
        JOIN employee_projects ep ON p.id = ep.project_id
        WHERE ep.employee_id = ?
        ORDER BY p.name
    ''', (current_user.id,)).fetchall()
    if not projects_list:
        projects_list = db.execute("SELECT id, name FROM projects ORDER BY name").fetchall()
    creators_list = db.execute('''
        SELECT DISTINCT u.id, u.full_name, u.username 
        FROM users u 
        JOIN tasks t ON t.creator_id = u.id 
        WHERE t.employee_id = ? 
        ORDER BY u.full_name
    ''', (current_user.id,)).fetchall()
    
    # Fetch options for filters specifically for subordinate tasks
    sub_creators_list = db.execute('''
        SELECT DISTINCT creator.id, creator.full_name, creator.username
        FROM users creator
        JOIN tasks t ON t.creator_id = creator.id
        JOIN employee_managers em ON t.employee_id = em.employee_id
        WHERE em.manager_id = ?
        ORDER BY creator.full_name
    ''', (current_user.id,)).fetchall()
    
    # Fetch assigned projects and remaining hour balances
    balance_rows = db.execute('''
        SELECT p.id, p.name, ep.allocated_hours,
               COALESCE(SUM(t.running_duration), 0) AS worked_seconds
        FROM projects p
        JOIN employee_projects ep ON p.id = ep.project_id
        LEFT JOIN tasks t ON p.id = t.project_id AND t.employee_id = ep.employee_id
        WHERE ep.employee_id = ?
        GROUP BY p.id
        ORDER BY p.name
    ''', (current_user.id,)).fetchall()
    
    assigned_project_balances = []
    for r in balance_rows:
        allocated = r['allocated_hours']
        worked_hours = r['worked_seconds'] / 3600.0
        remaining_hours = allocated - worked_hours if allocated > 0 else 0
        assigned_project_balances.append({
            'id': r['id'],
            'name': r['name'],
            'allocated_hours': allocated,
            'remaining_hours': remaining_hours
        })

    # Fetch project time split for charts
    chart_projects_labels = [r['name'] for r in balance_rows if r['worked_seconds'] > 0]
    chart_projects_data = [round(r['worked_seconds'] / 3600.0, 2) for r in balance_rows if r['worked_seconds'] > 0]
    
    # Fetch weekly daily history (last 7 days)
    chart_weekly_labels = []
    chart_weekly_data = []
    today = datetime.now().date()
    for i in range(6, -1, -1):
        day = today - timedelta(days=i)
        day_str = day.strftime('%Y-%m-%d')
        chart_weekly_labels.append(day.strftime('%b %d') if session.get('lang') != 'ar' else day.strftime('%m/%d'))
        
        row = db.execute('''
            SELECT COALESCE(SUM(running_duration), 0) AS daily_seconds
            FROM tasks
            WHERE employee_id = ? AND substr(created_at, 1, 10) = ?
        ''', (current_user.id, day_str)).fetchone()
        chart_weekly_data.append(round(row['daily_seconds'] / 3600.0, 2))
        
    # Fetch pending subordinate approvals explicitly routed to this manager
    pending_sub_approvals = db.execute('''
        SELECT t.*, u.full_name AS employee_name, p.name AS project_name
        FROM tasks t
        JOIN users u ON t.employee_id = u.id
        LEFT JOIN projects p ON t.project_id = p.id
        WHERE t.approver_id = ? AND t.approval_status = 'Submitted'
        ORDER BY t.created_at DESC
    ''', (current_user.id,)).fetchall()

    # Fetch this employee's list of managers
    my_managers = db.execute('''
        SELECT u.id, u.full_name, u.username
        FROM users u
        JOIN employee_managers em ON u.id = em.manager_id
        WHERE em.employee_id = ?
    ''', (current_user.id,)).fetchall()

    return render_template(
        'employee.html', 
        tasks=tasks, 
        subordinates=subordinates, 
        subordinate_tasks=subordinate_tasks,
        projects_list=projects_list,
        creators_list=creators_list,
        selected_projects=selected_projects,
        selected_statuses=selected_statuses,
        created_from=created_from,
        created_to=created_to,
        selected_creators=selected_creators,
        
        sub_creators_list=sub_creators_list,
        selected_sub_employees=selected_sub_employees,
        selected_sub_projects=selected_sub_projects,
        selected_sub_statuses=selected_sub_statuses,
        sub_created_from=sub_created_from,
        sub_created_to=sub_created_to,
        selected_sub_creators=selected_sub_creators,
        assigned_project_balances=assigned_project_balances,
        chart_projects_labels=chart_projects_labels,
        chart_projects_data=chart_projects_data,
        chart_weekly_labels=chart_weekly_labels,
        chart_weekly_data=chart_weekly_data,
        pending_sub_approvals=pending_sub_approvals,
        my_managers=my_managers
    )

@app.route('/employee/add', methods=['GET', 'POST'])
@role_required('Employee')
def add_task():
    db = get_db()
    if request.method == 'POST':
        title = request.form['title']
        description = request.form.get('description', '')
        project_id = int(request.form['project_id'])
        category = request.form.get('category', 'Other')
        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        db.execute('INSERT INTO tasks (employee_id, project_id, title, description, status, created_at, running_duration, creator_id, category) VALUES (?,?,?,?,?,?,?,?,?)',
                   (current_user.id, project_id, title, description, 'Pause', now, 0, current_user.id, category))
        db.commit()
        flash('Task added.', 'success')
        return redirect(url_for('employee_dashboard'))
    projects_raw = db.execute('''
        SELECT p.id, p.name, ep.allocated_hours,
               COALESCE(SUM(t.running_duration), 0) AS worked_seconds
        FROM projects p
        JOIN employee_projects ep ON p.id = ep.project_id
        LEFT JOIN tasks t ON p.id = t.project_id AND t.employee_id = ep.employee_id
        WHERE ep.employee_id = ?
        GROUP BY p.id
        ORDER BY p.name
    ''', (current_user.id,)).fetchall()
    
    projects = []
    for p in projects_raw:
        allocated = p['allocated_hours']
        if allocated == 0:
            name_with_balance = f"{p['name']} (Unlimited hours)"
        else:
            worked_hours = p['worked_seconds'] / 3600.0
            remaining = allocated - worked_hours
            name_with_balance = f"{p['name']} ({remaining:.2f} hrs remaining of {allocated} hrs)"
        projects.append({'id': p['id'], 'name': name_with_balance})
        
    if not projects:
        raw_fallback = db.execute('SELECT * FROM projects ORDER BY name').fetchall()
        projects = [{'id': p['id'], 'name': f"{p['name']} (Unlimited hours)"} for p in raw_fallback]
        
    return render_template('add_task.html', projects=projects)

@app.route('/employee/edit/<int:task_id>', methods=['GET', 'POST'])
@role_required('Employee')
def edit_task(task_id):
    db = get_db()
    task = db.execute('SELECT * FROM tasks WHERE id = ? AND employee_id = ?', (task_id, current_user.id)).fetchone()
    if not task:
        flash('Task not found.', 'danger')
        return redirect(url_for('employee_dashboard'))
    if task['creator_id'] != current_user.id:
        flash('You cannot edit tasks assigned by your manager.', 'danger')
        return redirect(url_for('employee_dashboard'))
    
    if request.method == 'POST':
        title = request.form['title']
        description = request.form['description']
        new_status = request.form['status']
        project_id = int(request.form['project_id'])
        category = request.form.get('category', 'Other')
        
        old_status = task['status']
        old_running_duration = task['running_duration'] or 0
        old_last_started_at = task['last_started_at']
        
        new_running_duration = old_running_duration
        new_last_started_at = old_last_started_at
        now_utc = datetime.utcnow()
        
        # Status change logic
        if old_status == 'Running' and new_status != 'Running':
            if old_last_started_at:
                try:
                    started_dt = datetime.fromisoformat(old_last_started_at)
                    elapsed = (now_utc - started_dt).total_seconds()
                    new_running_duration = old_running_duration + max(0, int(elapsed))
                except Exception:
                    pass
            new_last_started_at = None
        elif old_status != 'Running' and new_status == 'Running':
            new_last_started_at = now_utc.isoformat()
            
        approver_id = None
        approval_status = 'Draft'
        if new_status == 'Finish':
            if task['creator_id'] != current_user.id:
                approver_id = task['creator_id']
                approval_status = 'Submitted'
            else:
                req_approver = request.form.get('approver_id')
                if req_approver and req_approver != '' and req_approver != '0':
                    approver_id = int(req_approver)
                    approval_status = 'Submitted'
                else:
                    approver_id = None
                    approval_status = 'Approved'
        else:
            approval_status = 'Draft'
            approver_id = None

        db.execute('''
            UPDATE tasks 
            SET title = ?, description = ?, status = ?, running_duration = ?, last_started_at = ?, project_id = ?, category = ?, approval_status = ?, approver_id = ?
            WHERE id = ? AND employee_id = ?
        ''', (title, description, new_status, new_running_duration, new_last_started_at, project_id, category, approval_status, approver_id, task_id, current_user.id))
        db.commit()
        flash('Task updated.', 'success')
        return redirect(url_for('employee_dashboard'))
        
    projects_raw = db.execute('''
        SELECT p.id, p.name, ep.allocated_hours,
               COALESCE(SUM(t.running_duration), 0) AS worked_seconds
        FROM projects p
        JOIN employee_projects ep ON p.id = ep.project_id
        LEFT JOIN tasks t ON p.id = t.project_id AND t.employee_id = ep.employee_id
        WHERE ep.employee_id = ?
        GROUP BY p.id
        ORDER BY p.name
    ''', (current_user.id,)).fetchall()
    
    projects = []
    for p in projects_raw:
        allocated = p['allocated_hours']
        if allocated == 0:
            name_with_balance = f"{p['name']} (Unlimited hours)"
        else:
            worked_hours = p['worked_seconds'] / 3600.0
            remaining = allocated - worked_hours
            name_with_balance = f"{p['name']} ({remaining:.2f} hrs remaining of {allocated} hrs)"
        projects.append({'id': p['id'], 'name': name_with_balance})
        
    if not projects:
        raw_fallback = db.execute('SELECT * FROM projects ORDER BY name').fetchall()
        projects = [{'id': p['id'], 'name': f"{p['name']} (Unlimited hours)"} for p in raw_fallback]
        
    my_managers = db.execute('''
        SELECT u.id, u.full_name, u.username
        FROM users u
        JOIN employee_managers em ON u.id = em.manager_id
        WHERE em.employee_id = ?
    ''', (current_user.id,)).fetchall()

    return render_template('edit_task.html', task=task, projects=projects, my_managers=my_managers)

@app.route('/employee/task/<int:task_id>/update_status', methods=['POST'])
@role_required('Employee')
def update_task_status(task_id):
    db = get_db()
    task = db.execute('SELECT creator_id, employee_id, status, running_duration, last_started_at FROM tasks WHERE id = ? AND employee_id = ?', (task_id, current_user.id)).fetchone()
    if not task:
        flash('Task not found.', 'danger')
        return redirect(url_for('employee_dashboard'))
    new_status = request.form.get('status')
    valid_statuses = ['Running', 'Pause', 'Finish']
    if new_status in valid_statuses:
        old_status = task['status']
        old_running_duration = task['running_duration'] or 0
        old_last_started_at = task['last_started_at']
        
        new_running_duration = old_running_duration
        new_last_started_at = old_last_started_at
        now_utc = datetime.utcnow()
        
        # Status change logic
        if old_status == 'Running' and new_status != 'Running':
            if old_last_started_at:
                try:
                    started_dt = datetime.fromisoformat(old_last_started_at)
                    elapsed = (now_utc - started_dt).total_seconds()
                    new_running_duration = old_running_duration + max(0, int(elapsed))
                except Exception:
                    pass
            new_last_started_at = None
        elif old_status != 'Running' and new_status == 'Running':
            new_last_started_at = now_utc.isoformat()
            
        approver_id = None
        approval_status = 'Draft'
        if new_status == 'Finish':
            if task['creator_id'] != current_user.id:
                approver_id = task['creator_id']
                approval_status = 'Submitted'
            else:
                req_approver = request.form.get('approver_id')
                if req_approver and req_approver != '' and req_approver != '0':
                    approver_id = int(req_approver)
                    approval_status = 'Submitted'
                else:
                    approver_id = None
                    approval_status = 'Approved'
        else:
            approval_status = 'Draft'
            approver_id = None

        db.execute('''
            UPDATE tasks 
            SET status = ?, running_duration = ?, last_started_at = ?, approval_status = ?, approver_id = ?
            WHERE id = ? AND employee_id = ?
        ''', (new_status, new_running_duration, new_last_started_at, approval_status, approver_id, task_id, current_user.id))
        db.commit()
        flash('Task status updated.', 'success')
    else:
        flash('Invalid status.', 'danger')
    return redirect(url_for('employee_dashboard'))

@app.route('/employee/delete/<int:task_id>', methods=['POST'])
@role_required('Employee')
def delete_task(task_id):
    db = get_db()
    task = db.execute('SELECT creator_id, employee_id FROM tasks WHERE id = ?', (task_id,)).fetchone()
    if not task:
        flash('Task not found.', 'danger')
        return redirect(url_for('employee_dashboard'))
    if task['creator_id'] != current_user.id:
        flash('You can only delete tasks created by you.', 'danger')
        return redirect(url_for('employee_dashboard'))
    db.execute('DELETE FROM tasks WHERE id = ?', (task_id,))
    db.commit()
    flash('Task deleted.', 'success')
    if task['employee_id'] != current_user.id:
        return redirect(url_for('view_subordinate_tasks', sub_id=task['employee_id']))
    return redirect(url_for('employee_dashboard'))

@app.route('/employee/subordinates')
@role_required('Employee')
def view_subordinates():
    db = get_db()
    subordinates = db.execute('''
        SELECT u.* FROM users u
        JOIN employee_managers em ON u.id = em.employee_id
        WHERE em.manager_id = ?
    ''', (current_user.id,)).fetchall()
    return render_template('subordinates.html', subordinates=subordinates)

# ----- Subordinate Tasks -----
@app.route('/employee/subordinate/<int:sub_id>/tasks')
@role_required('Employee')
def view_subordinate_tasks(sub_id):
    db = get_db()
    link = db.execute('SELECT 1 FROM employee_managers WHERE employee_id = ? AND manager_id = ?', (sub_id, current_user.id)).fetchone()
    if not link:
        flash('Access denied.', 'danger')
        return redirect(url_for('employee_dashboard'))
    subordinate = db.execute('SELECT * FROM users WHERE id = ?', (sub_id,)).fetchone()
    tasks = db.execute('''
        SELECT t.*, u.full_name AS creator_name, p.name AS project_name
        FROM tasks t
        LEFT JOIN users u ON t.creator_id = u.id
        LEFT JOIN projects p ON t.project_id = p.id
        WHERE t.employee_id = ?
          AND EXISTS (
              SELECT 1 FROM employee_projects ep1
              WHERE ep1.employee_id = t.employee_id AND ep1.project_id = t.project_id
          )
          AND EXISTS (
              SELECT 1 FROM employee_projects ep2
              WHERE ep2.employee_id = ? AND ep2.project_id = t.project_id
          )
    ''', (sub_id, current_user.id)).fetchall()
    return render_template('subordinate_tasks.html', subordinate=subordinate, tasks=tasks)

@app.route('/employee/subordinate/<int:sub_id>/tasks/add', methods=['GET', 'POST'])
@role_required('Employee')
def add_subordinate_task(sub_id):
    db = get_db()
    link = db.execute('SELECT 1 FROM employee_managers WHERE employee_id = ? AND manager_id = ?', (sub_id, current_user.id)).fetchone()
    if not link:
        flash('Access denied.', 'danger')
        return redirect(url_for('employee_dashboard'))
    subordinate = db.execute('SELECT * FROM users WHERE id = ?', (sub_id,)).fetchone()
    if request.method == 'POST':
        title = request.form['title']
        description = request.form.get('description', '')
        project_id = int(request.form['project_id'])
        category = request.form.get('category', 'Other')
        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        db.execute('INSERT INTO tasks (employee_id, project_id, title, description, status, created_at, running_duration, creator_id, category) VALUES (?,?,?,?,?,?,?,?,?)',
                   (sub_id, project_id, title, description, 'Pause', now, 0, current_user.id, category))
        db.commit()
        flash('Task assigned to employee.', 'success')
        return redirect(url_for('view_subordinate_tasks', sub_id=sub_id))
    projects_raw = db.execute('''
        SELECT p.id, p.name, ep_sub.allocated_hours,
               COALESCE(SUM(t.running_duration), 0) AS worked_seconds
        FROM projects p
        JOIN employee_projects ep_sub ON p.id = ep_sub.project_id
        JOIN employee_projects ep_mgr ON p.id = ep_mgr.project_id
        LEFT JOIN tasks t ON p.id = t.project_id AND t.employee_id = ep_sub.employee_id
        WHERE ep_sub.employee_id = ? AND ep_mgr.employee_id = ?
        GROUP BY p.id
        ORDER BY p.name
    ''', (sub_id, current_user.id)).fetchall()
    
    if not projects_raw:
        projects_raw = db.execute('''
            SELECT p.id, p.name, ep.allocated_hours,
                   COALESCE(SUM(t.running_duration), 0) AS worked_seconds
            FROM projects p
            JOIN employee_projects ep ON p.id = ep.project_id
            LEFT JOIN tasks t ON p.id = t.project_id AND t.employee_id = ep.employee_id
            WHERE ep.employee_id = ?
            GROUP BY p.id
            ORDER BY p.name
        ''', (sub_id,)).fetchall()
        
    projects = []
    for p in projects_raw:
        allocated = p['allocated_hours']
        if allocated == 0:
            name_with_balance = f"{p['name']} (Unlimited hours)"
        else:
            worked_hours = p['worked_seconds'] / 3600.0
            remaining = allocated - worked_hours
            name_with_balance = f"{p['name']} ({remaining:.2f} hrs remaining of {allocated} hrs)"
        projects.append({'id': p['id'], 'name': name_with_balance})
        
    if not projects:
        raw_fallback = db.execute('SELECT * FROM projects ORDER BY name').fetchall()
        projects = [{'id': p['id'], 'name': f"{p['name']} (Unlimited hours)"} for p in raw_fallback]
        
    return render_template('add_subordinate_task.html', subordinate=subordinate, projects=projects)

# Removed bulk submit_timesheet endpoint. Individual submissions are handled dynamically.

@app.route('/tasks/approve/<int:task_id>', methods=['POST'])
@role_required('Employee')
def approve_task(task_id):
    db = get_db()
    task = db.execute('SELECT * FROM tasks WHERE id = ?', (task_id,)).fetchone()
    if not task:
        flash('Task not found.', 'danger')
        return redirect(request.referrer or url_for('index'))
        
    allowed = False
    link = db.execute('SELECT 1 FROM employee_managers WHERE employee_id = ? AND manager_id = ?', 
                      (task['employee_id'], current_user.id)).fetchone()
    if link:
        allowed = True
            
    if not allowed:
        flash('Permission denied.', 'danger')
        return redirect(request.referrer or url_for('index'))
        
    db.execute('UPDATE tasks SET approval_status = "Approved", approval_comments = NULL WHERE id = ?', (task_id,))
    db.commit()
    flash('Task timesheet approved.', 'success')
    return redirect(request.referrer or url_for('index'))

@app.route('/tasks/reject/<int:task_id>', methods=['POST'])
@role_required('Employee')
def reject_task(task_id):
    db = get_db()
    task = db.execute('SELECT * FROM tasks WHERE id = ?', (task_id,)).fetchone()
    if not task:
        flash('Task not found.', 'danger')
        return redirect(request.referrer or url_for('index'))
        
    allowed = False
    link = db.execute('SELECT 1 FROM employee_managers WHERE employee_id = ? AND manager_id = ?', 
                      (task['employee_id'], current_user.id)).fetchone()
    if link:
        allowed = True
            
    if not allowed:
        flash('Permission denied.', 'danger')
        return redirect(request.referrer or url_for('index'))
        
    comments = request.form.get('comments', '')
    db.execute('UPDATE tasks SET approval_status = "Rejected", approval_comments = ? WHERE id = ?', (comments, task_id))
    db.commit()
    flash('Task timesheet rejected.', 'warning')
    return redirect(request.referrer or url_for('index'))

@app.route('/hr/reports/print')
@role_required('HR')
def hr_reports_print():
    report_type = request.args.get('type')
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    category = request.args.get('category')
    
    if not report_type or not date_from or not date_to:
        flash('Missing report parameters.', 'danger')
        return redirect(url_for('hr_reports_page'))
        
    db = get_db()
    rows = []
    title_label = ""
    
    if report_type == 'employee':
        employee_id = request.args.get('employee_id')
        if not employee_id:
            flash('Employee required.', 'danger')
            return redirect(url_for('hr_reports_page'))
        emp = db.execute('SELECT full_name FROM users WHERE id = ?', (employee_id,)).fetchone()
        if not emp:
            flash('Employee not found.', 'danger')
            return redirect(url_for('hr_reports_page'))
        title_label = f"Employee Report: {emp['full_name']}"
        project_ids = request.args.getlist('project_ids')
        rows = _fetch_tasks_for_employee_report(employee_id, date_from, date_to, project_ids, category)
    elif report_type == 'project':
        project_id = request.args.get('project_id')
        if project_id and project_id != '':
            proj = db.execute('SELECT name FROM projects WHERE id = ?', (project_id,)).fetchone()
            if not proj:
                flash('Project not found.', 'danger')
                return redirect(url_for('hr_reports_page'))
            title_label = f"Project Report: {proj['name']}"
            rows = _fetch_tasks_for_project_report(project_id, date_from, date_to, category=category)
        else:
            title_label = "Project Report: All Projects"
            rows = _fetch_tasks_for_project_report(None, date_from, date_to, category=category)
            
    total_seconds_sum = sum(r['running_duration'] for r in rows)
    total_hours_sum = round(total_seconds_sum / 3600.0, 2)
    
    return render_template(
        'print_report.html',
        rows=rows,
        title_label=title_label,
        date_from=date_from,
        date_to=date_to,
        category=category,
        total_hours=total_hours_sum,
        generated_by=current_user.full_name or current_user.username,
        generated_on=datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    )

# ---------- Init ----------
if __name__ == '__main__':
    with app.app_context():
        init_db()
        seed_data()
    app.run(host='0.0.0.0', port=5000, debug=True)
